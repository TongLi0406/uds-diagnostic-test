#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
UDS诊断调查表解析器
支持解析Excel(.xlsx/.xls), CSV, JSON格式的诊断调查表
输出标准化JSON用于测试脚本生成
"""

__version__ = "1.6.0"

import argparse
import json
import os
import re
import sys
from pathlib import Path


# ============================================================================
# 默认值定义
# ============================================================================
DEFAULTS = {
    "did": {
        "did_name": "Unknown DID",
        "did_name_cn": "未知DID",
        "rw_state": "R",
        "size_bytes": 1,
        "data_type": "RAW",
        "method_type": "identical",
        "range_min": None,  # will be computed from size
        "range_max": None,
        "default_value": 0,
        "unit": "- (No Unit)",
        "storage_pos": "Unknown",
        "functional_addressing": "N",
        "session_default": "Y",
        "session_programming": "N",
        "session_extended": "Y",
        "session_boot_default": "N",
        "session_boot_programming": "N",
        "session_boot_extended": "N",
        "security_level0": "Y",
        "security_level1": "N",
        "security_fbl": "N",
        "security_immo": "N",
        "write_session_default": "N",
        "write_session_programming": "N",
        "write_session_extended": "N",
        "write_security_level0": "N",
        "write_security_level1": "N",
        "write_security_fbl": "N",
        "write_security_immo": "N",
    },
    "io_control": {
        "io_param": "ShortTermAdjustment",
        "conditions": "Extended Session",
        "security_level": "Level1",
    },
    "routine": {
        "control_type": "01",
        "conditions": "Extended Session",
        "security_level": "Level1",
    },
    "dtc": {
        "failure_type": "0x00",
        "priority": 4,
        "lamp_flag": "unused",
        "monitor_type": "continuous",
        "monitor_rate": "100ms",
        "dtc_aging": 40,
    },
}


def safe_str(val):
    """安全转换为字符串"""
    if val is None:
        return ""
    return str(val).strip()


def safe_int(val, default=0):
    """安全转换为整数"""
    if val is None:
        return default
    try:
        s = str(val).strip()
        if s.startswith("0x") or s.startswith("0X"):
            return int(s, 16)
        return int(float(s))
    except (ValueError, TypeError):
        return default


def normalize_yn(val):
    """标准化Y/N值"""
    s = safe_str(val).upper()
    if s in ("Y", "YES", "是", "TRUE", "1"):
        return "Y"
    return "N"


def normalize_hex(val):
    """标准化HEX值"""
    s = safe_str(val).upper().strip()
    if not s or s == "/":
        return None
    s = s.replace(" ", "")
    if s.startswith("0X"):
        return "0x" + s[2:]
    if re.match(r'^[0-9A-F]+$', s):
        return "0x" + s
    return s


def normalize_can_id(val):
    """严格标准化CAN ID，非法描述文本返回None。
    规则:
    - 带0x前缀: 直接按十六进制解析
    - 含A-F字母的纯hex串: 按十六进制解析
    - 纯数字: 仅当>=0x100(256)时才接受, 防止把"2016"等年份/行号误判为CAN ID
    - 合法范围: [0, 0x1FFFFFFF]
    """
    s = safe_str(val).strip()
    if not s or s == "/":
        return None

    compact = s.replace(" ", "").replace("_", "")
    if compact.lower().startswith("0x"):
        digits = compact[2:]
        base = 16
    elif re.fullmatch(r'[0-9A-Fa-f]+', compact) and re.search(r'[A-Fa-f]', compact):
        digits = compact
        base = 16
    elif re.fullmatch(r'\d+', compact):
        digits = compact
        base = 10
    else:
        return None

    try:
        can_id = int(digits, base)
    except ValueError:
        return None

    # 纯数字(无0x/无hex字母)时, 排除过小的值(可能是行号/年份)
    if base == 10 and not compact.lower().startswith("0x") and can_id < 0x100:
        return None

    if 0 <= can_id <= 0x1FFFFFFF:
        return f"0x{can_id:X}"
    return None


def iter_nearby_values(ws, row_idx, col_idx, max_cols=6, max_rows=2):
    """从标签右侧及下方邻近区域枚举候选值。"""
    yielded = set()

    for c in range(col_idx + 1, min(ws.max_column, col_idx + max_cols) + 1):
        raw = safe_str(ws.cell(row=row_idx, column=c).value).strip()
        if raw and raw not in yielded:
            yielded.add(raw)
            yield raw

    for r in range(row_idx + 1, min(ws.max_row, row_idx + max_rows) + 1):
        start_col = max(1, col_idx - 1)
        end_col = min(ws.max_column, col_idx + max_cols)
        for c in range(start_col, end_col + 1):
            raw = safe_str(ws.cell(row=r, column=c).value).strip()
            if raw and raw not in yielded:
                yielded.add(raw)
                yield raw


def detect_column_mapping(header_row, sheet_type="did"):
    """
    根据表头行自动检测列映射关系
    返回 {标准字段名: 列索引} 的映射
    """
    mapping = {}
    headers = {}
    for idx, cell in enumerate(header_row):
        val = safe_str(cell).lower()
        headers[idx] = val

    if sheet_type == "did":
        patterns = {
            "num": r"num|序号|no\.|#",
            "did_number": r"did.*(hex|number|号)|did\b",
            "did_name": r"did.*description.*english|did.*name(?!.*chin)",
            "did_name_cn": r"did.*description.*chinese|did.*name.*chin|did名称.*中",
            "cvt": r"cvt|约定值",
            "rw_state": r"r/w|rw.*state|读.*写|读写",
            "size_bytes": r"size.*byte|字节数|长度",
            "byte_pos": r"\bbyte\b|字节(?!数)",
            "bit_pos": r"\bbit\b|位(?!数)",
            "sub_data_name": r"sub.*data.*name.*english|子数据.*英|parameter(?!.*chin)",
            "sub_data_name_cn": r"sub.*data.*name.*chinese|子数据.*中|parameter.*chin",
            "range_min": r"range.*min|最小值|min.*phy",
            "range_max": r"range.*max|最大值|max.*phy",
            "unit": r"\bunit\b|单位",
            "method_type": r"method.*type.*english|方法类型.*英|methodtype(?!.*chin)",
            "method_type_cn": r"method.*type.*chinese|方法类型.*中",
            "default_value": r"default.*value|默认值",
            "data_type": r"data.*type|数据类型|datatype",
            "storage_pos": r"storage.*pos|存储位置",
            "functional_addressing": r"functional|功能寻址",
            "remarks": r"remark|备注|comment",
            # Session/Security columns detected by position relative to service headers
        }
    elif sheet_type == "io_control":
        patterns = {
            "io_param": r"iocontrol.*param|io参数",
            "did_number": r"did.*(hex|number)|did\b",
            "did_name": r"did.*name(?!.*chin)|io.*id.*name(?!.*chin)",
            "did_name_cn": r"did.*name.*chin|io.*id.*name.*chin|io.*name.*中",
            "req_resp": r"req.*resp|请求.*响应",
            "parameter": r"parameter(?!.*chin)",
            "byte_pos": r"byte.*pos|字节",
            "bit_pos": r"bit.*pos|位",
            "bit_length": r"bit.*length|位长",
            "data_type": r"data.*type|数据类型",
            "method_type": r"method.*type|方法类型",
            "unit": r"\bunit\b|单位",
            "security_level": r"security.*level|安全.*等级",
            "conditions": r"condition|条件",
            "remarks": r"remark|备注|comment",
        }
    elif sheet_type == "routine":
        patterns = {
            "control_type": r"control.*type|控制类型",
            "rid_number": r"rid.*(hex|number)|rid\b",
            "rid_name": r"rid.*name(?!.*chin)",
            "rid_name_cn": r"rid.*name.*chin|rid.*名称.*中",
            "req_resp": r"req.*resp|请求.*响应",
            "parameter": r"parameter(?!.*chin)",
            "byte_pos": r"byte.*pos|字节",
            "bit_pos": r"bit.*pos|位",
            "bit_length": r"bit.*length|位长",
            "data_type": r"data.*type|数据类型",
            "method_type": r"method.*type|方法类型",
            "unit": r"\bunit\b|单位",
            "security_level": r"security.*level|安全.*等级",
            "conditions": r"condition|条件",
            "remarks": r"remark|备注|comment",
        }
    elif sheet_type == "dtc":
        patterns = {
            "num": r"num|序号|no\.",
            "dtc_number_hex": r"dtc.*number.*hex|dtc.*hex",
            "dtc_number": r"dtc.*number(?!.*hex)|dtc编号",
            "failure_type": r"failure.*type|故障类型",
            "dtc_name": r"dtc.*name|故障名称|故障描述",
            "priority": r"priority|优先级",
            "lamp_flag": r"lamp.*flag|灯标志",
            "monitor_enable": r"monitor.*enable|监测.*使能|监测.*条件",
            "monitor_type": r"monitor.*type|监测.*类型",
            "monitor_rate": r"monitor.*rate|监测.*频率",
            "test_failed_criteria": r"test.*fail|判定.*fail|confirmation.*fail|成熟.*条件",
            "mature_time": r"mature.*time|成熟时间",
            "test_pass_criteria": r"test.*pass|判定.*pass|恢复条件",
            "demature_time": r"de.*mature|去成熟|恢复时间",
            "ecu_action": r"ecu.*fail.*action|limp|故障动作",
            "dtc_aging": r"dtc.*aging|老化",
            "snapshot": r"snapshot|快照",
            "remarks": r"remark|备注|comment",
        }
    else:
        patterns = {}

    for field_name, pattern in patterns.items():
        for idx, header_text in headers.items():
            if re.search(pattern, header_text, re.IGNORECASE):
                if field_name not in mapping:
                    mapping[field_name] = idx
                break

    return mapping


def detect_session_security_columns(ws, header_row_idx):
    """
    检测会话模式和安全访问等级的列位置
    返回 read_sessions, read_security, write_sessions, write_security 的列映射
    """
    result = {
        "read_sessions": {},    # {session_name: col_idx}
        "read_security": {},    # {level_name: col_idx}
        "write_sessions": {},
        "write_security": {},
    }

    # 查找Service $22 和 $2E 的起始列
    service_22_col = None
    service_2e_col = None

    for row in ws.iter_rows(min_row=max(1, header_row_idx - 2),
                            max_row=header_row_idx,
                            values_only=False):
        for cell in row:
            val = safe_str(cell.value).lower()
            if "$22" in val or "service 22" in val or "0x22" in val:
                service_22_col = cell.column - 1
            elif "$2e" in val or "service 2e" in val or "0x2e" in val:
                service_2e_col = cell.column - 1

    # 查找会话和安全列
    # 扫描header行附近几行来找session和security子标题
    for row in ws.iter_rows(min_row=max(1, header_row_idx - 1),
                            max_row=header_row_idx + 2,
                            values_only=False):
        for cell in row:
            val = safe_str(cell.value).lower()
            col = cell.column - 1

            session_map = None
            security_map = None

            # 判断属于$22还是$2E区域
            if service_22_col is not None and service_2e_col is not None:
                if col < service_2e_col:
                    session_map = result["read_sessions"]
                    security_map = result["read_security"]
                else:
                    session_map = result["write_sessions"]
                    security_map = result["write_security"]
            elif service_22_col is not None:
                session_map = result["read_sessions"]
                security_map = result["read_security"]

            if session_map is None:
                session_map = result["read_sessions"]
                security_map = result["read_security"]

            # 识别会话模式
            if "0x01" in val and ("默认" in val or "default" in val):
                session_map["default_0x01"] = col
            elif "0x02" in val and ("编程" in val or "program" in val):
                session_map["programming_0x02"] = col
            elif "0x03" in val and ("扩展" in val or "extend" in val):
                session_map["extended_0x03"] = col

            # 识别安全等级
            if "level0" in val or ("锁定" in val and "level" in val):
                security_map["level0_locked"] = col
            elif "level1" in val or ("1级" in val and "解锁" in val):
                security_map["level1"] = col
            elif "levelfbl" in val or "fbl" in val.replace(" ", ""):
                security_map["level_fbl"] = col
            elif "levelimmo" in val or "immo" in val:
                security_map["level_immo"] = col

    # 简化格式兜底: Application / Boot Loader / Secturity Level 列 (ECU DID等sheet)
    # 这些sheet没有$22/$2E子标题，而是直接用Application列标记 "R"/"R/W"/"W"
    if not result["read_sessions"]:
        alt_app_col = None
        alt_boot_col = None
        alt_sec_col = None
        for row in ws.iter_rows(min_row=max(1, header_row_idx - 1),
                                max_row=header_row_idx + 1,
                                values_only=False):
            for cell in row:
                val = safe_str(cell.value).lower().strip()
                col = cell.column - 1
                if val in ("application", "应用") or "application" in val:
                    alt_app_col = col
                elif "boot" in val and "loader" in val:
                    alt_boot_col = col
                elif "security" in val or "secturity" in val:
                    alt_sec_col = col

        if alt_app_col is not None or alt_sec_col is not None:
            result["_alt_format"] = {
                "application_col": alt_app_col,
                "boot_col": alt_boot_col,
                "security_col": alt_sec_col,
            }

    return result


def detect_sheet_type(sheet_name, ws):
    """根据Sheet名称和内容检测诊断服务类型"""
    name_lower = sheet_name.lower()

    # bak检查必须在did之前——否则 "System DID_bak" 会先匹配did
    if "bak" in name_lower:
        return None  # 跳过备份sheet
    elif "dtc" in name_lower:
        return "dtc"
    elif "iocontrol" in name_lower or "0x2f" in name_lower or "2f" in name_lower:
        return "io_control"
    elif "routine" in name_lower or "0x31" in name_lower:
        return "routine"
    elif "did" in name_lower:
        return "did"

    # 通过内容检测
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        row_text = " ".join([safe_str(c) for c in row]).lower()
        if "did" in row_text and ("read" in row_text or "write" in row_text or "r/w" in row_text):
            return "did"
        elif "iocontrol" in row_text or "io control" in row_text:
            return "io_control"
        elif "routine" in row_text or "rid" in row_text:
            return "routine"
        elif "dtc" in row_text and "fault" in row_text:
            return "dtc"

    return None


def find_header_row(ws, sheet_type):
    """找到表头行的行号（1-based）"""
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        row_text = " ".join([safe_str(c) for c in row if c is not None]).lower()

        if sheet_type == "did":
            if ("did" in row_text and ("number" in row_text or "hex" in row_text)) or \
               ("did" in row_text and "name" in row_text and "r/w" in row_text):
                return row_idx
        elif sheet_type == "io_control":
            if "iocontrol" in row_text or ("did" in row_text and "parameter" in row_text):
                return row_idx
        elif sheet_type == "routine":
            if "rid" in row_text or ("control" in row_text and "routine" in row_text):
                return row_idx
        elif sheet_type == "dtc":
            if "dtc" in row_text and ("number" in row_text or "name" in row_text):
                return row_idx

    return 1  # 默认第一行


def parse_did_sheet(ws, sheet_name):
    """解析DID Sheet"""
    dids = []
    defaults_used = []

    header_row_idx = find_header_row(ws, "did")
    header_row = [safe_str(cell.value) for cell in list(ws.iter_rows(
        min_row=header_row_idx, max_row=header_row_idx, values_only=False))[0]]

    col_map = detect_column_mapping(header_row, "did")
    session_cols = detect_session_security_columns(ws, header_row_idx)

    current_did = None

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=False):
        cells = [cell.value for cell in row]

        # 检查结束标记
        if any(safe_str(c).startswith("#End") for c in cells if c):
            break

        # 跳过完全空行
        if all(c is None for c in cells):
            continue

        # 检查是不是新DID还是子数据行
        did_num_idx = col_map.get("did_number")
        did_num_val = cells[did_num_idx] if did_num_idx is not None and did_num_idx < len(cells) else None

        if did_num_val is not None and safe_str(did_num_val).strip():
            # 新DID
            did_hex = normalize_hex(did_num_val)
            if did_hex is None:
                continue

            did_entry = {
                "did_number": did_hex,
                "did_name": "",
                "did_name_cn": "",
                "rw_state": DEFAULTS["did"]["rw_state"],
                "size_bytes": DEFAULTS["did"]["size_bytes"],
                "data_type": DEFAULTS["did"]["data_type"],
                "method_type": DEFAULTS["did"]["method_type"],
                "unit": DEFAULTS["did"]["unit"],
                "default_value": DEFAULTS["did"]["default_value"],
                "storage_pos": DEFAULTS["did"]["storage_pos"],
                "functional_addressing": DEFAULTS["did"]["functional_addressing"],
                "range_min": None,
                "range_max": None,
                "remarks": "",
                "sub_data": [],
                "read_sessions": {"default_0x01": "N", "programming_0x02": "N", "extended_0x03": "N"},
                "read_security": {"level0_locked": "Y", "level1": "N", "level_fbl": "N", "level_immo": "N"},
                "write_sessions": {"default_0x01": "N", "programming_0x02": "N", "extended_0x03": "N"},
                "write_security": {"level0_locked": "N", "level1": "N", "level_fbl": "N", "level_immo": "N"},
                "boot_read_sessions": {"default_0x01": "N", "programming_0x02": "N", "extended_0x03": "N"},
                "boot_write_sessions": {"default_0x01": "N", "programming_0x02": "N", "extended_0x03": "N"},
            }

            # 填充各字段
            field_map = {
                "did_name": "did_name", "did_name_cn": "did_name_cn",
                "rw_state": "rw_state", "data_type": "data_type",
                "method_type": "method_type", "unit": "unit",
                "default_value": "default_value", "storage_pos": "storage_pos",
                "functional_addressing": "functional_addressing",
                "remarks": "remarks",
            }

            for field, key in field_map.items():
                idx = col_map.get(field)
                if idx is not None and idx < len(cells) and cells[idx] is not None:
                    val = safe_str(cells[idx])
                    if val and val != "/":
                        did_entry[key] = val

            # Size
            size_idx = col_map.get("size_bytes")
            if size_idx is not None and size_idx < len(cells) and cells[size_idx] is not None:
                did_entry["size_bytes"] = safe_int(cells[size_idx], DEFAULTS["did"]["size_bytes"])
            # size_bytes=0 无效，回退为默认值1
            if did_entry["size_bytes"] <= 0:
                did_entry["size_bytes"] = DEFAULTS["did"]["size_bytes"]

            # Range (解析为数值，支持hex/decimal)
            min_idx = col_map.get("range_min")
            max_idx = col_map.get("range_max")
            if min_idx is not None and min_idx < len(cells) and cells[min_idx] is not None:
                min_val = safe_str(cells[min_idx])
                if min_val and min_val != "/":
                    did_entry["range_min"] = safe_int(cells[min_idx], None)
                    if did_entry["range_min"] is None:
                        # 尝试保留字符串（可能是公式描述）
                        did_entry["range_min"] = min_val
            if max_idx is not None and max_idx < len(cells) and cells[max_idx] is not None:
                max_val = safe_str(cells[max_idx])
                if max_val and max_val != "/":
                    did_entry["range_max"] = safe_int(cells[max_idx], None)
                    if did_entry["range_max"] is None:
                        did_entry["range_max"] = max_val

            # Functional addressing
            fa_idx = col_map.get("functional_addressing")
            if fa_idx is not None and fa_idx < len(cells):
                did_entry["functional_addressing"] = normalize_yn(cells[fa_idx])

            # Session & Security from detected columns
            for session_key, col_idx in session_cols.get("read_sessions", {}).items():
                if col_idx < len(cells):
                    did_entry["read_sessions"][session_key] = normalize_yn(cells[col_idx])

            for sec_key, col_idx in session_cols.get("read_security", {}).items():
                if col_idx < len(cells):
                    did_entry["read_security"][sec_key] = normalize_yn(cells[col_idx])

            for session_key, col_idx in session_cols.get("write_sessions", {}).items():
                if col_idx < len(cells):
                    did_entry["write_sessions"][session_key] = normalize_yn(cells[col_idx])

            for sec_key, col_idx in session_cols.get("write_security", {}).items():
                if col_idx < len(cells):
                    did_entry["write_security"][sec_key] = normalize_yn(cells[col_idx])

            # 简化格式处理 (ECU DID等sheet: Application/Boot Loader/Security Level列)
            alt_fmt = session_cols.get("_alt_format")
            if alt_fmt:
                app_col = alt_fmt.get("application_col")
                boot_col = alt_fmt.get("boot_col")
                sec_col = alt_fmt.get("security_col")

                # Application列: "R" → 在default+extended可读, "R/W" → 还可写, "W" → 仅写
                if app_col is not None and app_col < len(cells) and cells[app_col]:
                    app_val = safe_str(cells[app_col]).upper().strip()
                    if "R" in app_val:
                        did_entry["read_sessions"]["default_0x01"] = "Y"
                        did_entry["read_sessions"]["extended_0x03"] = "Y"
                    if "W" in app_val:
                        did_entry["write_sessions"]["extended_0x03"] = "Y"

                # Boot Loader列: "R" → programming可读, "R/W" → 也可写
                if boot_col is not None and boot_col < len(cells) and cells[boot_col]:
                    boot_val = safe_str(cells[boot_col]).upper().strip()
                    if boot_val != "N/A" and boot_val != "-":
                        if "R" in boot_val:
                            did_entry["read_sessions"]["programming_0x02"] = "Y"
                        if "W" in boot_val:
                            did_entry["write_sessions"]["programming_0x02"] = "Y"

                # Security Level列: "R/L0" "R/L0 W/L1" "R/L0,L1" 等
                if sec_col is not None and sec_col < len(cells) and cells[sec_col]:
                    sec_val = safe_str(cells[sec_col]).upper().strip()
                    # 解析读取安全: R/L0 表示读取需要Level0(锁定)
                    if "R/" in sec_val or "R /" in sec_val:
                        r_part = sec_val.split("W")[0] if "W" in sec_val else sec_val
                        if "L0" in r_part:
                            did_entry["read_security"]["level0_locked"] = "Y"
                        if "L1" in r_part:
                            did_entry["read_security"]["level1"] = "Y"
                        if "L3" in r_part or "FBL" in r_part:
                            did_entry["read_security"]["level_fbl"] = "Y"
                        if "L5" in r_part or "IMMO" in r_part:
                            did_entry["read_security"]["level_immo"] = "Y"
                    # 解析写入安全: W/L1 表示写入需要Level1
                    if "W/" in sec_val or "W /" in sec_val:
                        w_part = sec_val[sec_val.index("W"):]
                        if "L0" in w_part:
                            did_entry["write_security"]["level0_locked"] = "Y"
                        if "L1" in w_part:
                            did_entry["write_security"]["level1"] = "Y"
                        if "L3" in w_part or "FBL" in w_part:
                            did_entry["write_security"]["level_fbl"] = "Y"
                        if "L5" in w_part or "IMMO" in w_part:
                            did_entry["write_security"]["level_immo"] = "Y"

            # 记录使用了默认值的字段
            did_defaults = []
            if not did_entry["did_name"] or did_entry["did_name"] == DEFAULTS["did"]["did_name"]:
                did_entry["did_name"] = DEFAULTS["did"]["did_name"]
                did_defaults.append(("did_name", DEFAULTS["did"]["did_name"]))
            if did_entry["rw_state"] == DEFAULTS["did"]["rw_state"] and \
               (col_map.get("rw_state") is None or cells[col_map["rw_state"]] is None):
                did_defaults.append(("rw_state", DEFAULTS["did"]["rw_state"]))
            if did_entry["range_min"] is None:
                did_entry["range_min"] = 0
                did_defaults.append(("range_min", "0x00 (基于数据类型)"))
            if did_entry["range_max"] is None:
                did_entry["range_max"] = (1 << (did_entry["size_bytes"] * 8)) - 1
                did_defaults.append(("range_max", f"0x{'FF' * did_entry['size_bytes']} (基于Size)"))

            if did_defaults:
                defaults_used.append({
                    "did": did_hex,
                    "name": did_entry["did_name"],
                    "defaults": did_defaults,
                })

            # 解析子数据（当前行）
            byte_idx = col_map.get("byte_pos")
            bit_idx = col_map.get("bit_pos")
            sub_name_idx = col_map.get("sub_data_name")
            sub_name_cn_idx = col_map.get("sub_data_name_cn")

            if byte_idx is not None and cells[byte_idx] is not None:
                sub = {
                    "byte_pos": safe_str(cells[byte_idx]),
                    "bit_pos": safe_str(cells[bit_idx]) if bit_idx and bit_idx < len(cells) else "ALL",
                    "sub_name": safe_str(cells[sub_name_idx]) if sub_name_idx and sub_name_idx < len(cells) else "",
                    "sub_name_cn": safe_str(cells[sub_name_cn_idx]) if sub_name_cn_idx and sub_name_cn_idx < len(cells) else "",
                    "data_type": did_entry["data_type"],
                    "method_type": did_entry["method_type"],
                    "range_min": did_entry["range_min"],
                    "range_max": did_entry["range_max"],
                }
                did_entry["sub_data"].append(sub)

            current_did = did_entry
            dids.append(did_entry)

        elif current_did is not None:
            # 子数据行（DID编号为空但有数据的行）
            byte_idx = col_map.get("byte_pos")
            if byte_idx is not None and byte_idx < len(cells) and cells[byte_idx] is not None:
                sub_name_idx = col_map.get("sub_data_name")
                sub_name_cn_idx = col_map.get("sub_data_name_cn")
                dt_idx = col_map.get("data_type")
                mt_idx = col_map.get("method_type")
                min_idx = col_map.get("range_min")
                max_idx = col_map.get("range_max")
                bit_idx = col_map.get("bit_pos")

                sub = {
                    "byte_pos": safe_str(cells[byte_idx]),
                    "bit_pos": safe_str(cells[bit_idx]) if bit_idx and bit_idx < len(cells) and cells[bit_idx] else "ALL",
                    "sub_name": safe_str(cells[sub_name_idx]) if sub_name_idx and sub_name_idx < len(cells) else "",
                    "sub_name_cn": safe_str(cells[sub_name_cn_idx]) if sub_name_cn_idx and sub_name_cn_idx < len(cells) else "",
                    "data_type": safe_str(cells[dt_idx]) if dt_idx and dt_idx < len(cells) and cells[dt_idx] else current_did["data_type"],
                    "method_type": safe_str(cells[mt_idx]) if mt_idx and mt_idx < len(cells) and cells[mt_idx] else current_did["method_type"],
                    "range_min": safe_str(cells[min_idx]) if min_idx and min_idx < len(cells) and cells[min_idx] else None,
                    "range_max": safe_str(cells[max_idx]) if max_idx and max_idx < len(cells) and cells[max_idx] else None,
                }
                current_did["sub_data"].append(sub)

    return dids, defaults_used


def parse_io_control_sheet(ws, sheet_name):
    """解析IOControl Sheet"""
    io_controls = []
    defaults_used = []

    header_row_idx = find_header_row(ws, "io_control")
    header_row = [safe_str(cell.value) for cell in list(ws.iter_rows(
        min_row=header_row_idx, max_row=header_row_idx, values_only=False))[0]]

    col_map = detect_column_mapping(header_row, "io_control")

    current_io = None

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=False):
        cells = [cell.value for cell in row]
        if any(safe_str(c).startswith("#End") for c in cells if c):
            break
        if all(c is None for c in cells):
            continue

        did_idx = col_map.get("did_number")
        did_val = cells[did_idx] if did_idx is not None and did_idx < len(cells) else None

        if did_val is not None and safe_str(did_val).strip():
            did_hex = normalize_hex(did_val)
            if did_hex is None:
                continue

            io_entry = {
                "did_number": did_hex,
                "did_name": "",
                "did_name_cn": "",
                "io_param": DEFAULTS["io_control"]["io_param"],
                "conditions": DEFAULTS["io_control"]["conditions"],
                "security_level": DEFAULTS["io_control"]["security_level"],
                "parameters": [],
                "remarks": "",
            }

            for field in ["did_name", "did_name_cn", "io_param", "conditions",
                          "security_level", "remarks"]:
                idx = col_map.get(field)
                if idx is not None and idx < len(cells) and cells[idx] is not None:
                    io_entry[field] = safe_str(cells[idx])

            io_defaults = []
            if not io_entry["conditions"] or io_entry["conditions"] == DEFAULTS["io_control"]["conditions"]:
                if col_map.get("conditions") is None:
                    io_defaults.append(("conditions", DEFAULTS["io_control"]["conditions"]))
            if not io_entry["security_level"] or io_entry["security_level"] == DEFAULTS["io_control"]["security_level"]:
                if col_map.get("security_level") is None:
                    io_defaults.append(("security_level", DEFAULTS["io_control"]["security_level"]))

            if io_defaults:
                defaults_used.append({"io_did": did_hex, "defaults": io_defaults})

            # 解析参数
            req_resp_idx = col_map.get("req_resp")
            param_idx = col_map.get("parameter")
            byte_idx = col_map.get("byte_pos")
            bit_idx = col_map.get("bit_pos")
            bitlen_idx = col_map.get("bit_length")
            dt_idx = col_map.get("data_type")

            if param_idx is not None and param_idx < len(cells) and cells[param_idx] is not None:
                param = {
                    "req_resp": safe_str(cells[req_resp_idx]) if req_resp_idx and req_resp_idx < len(cells) else "Req",
                    "name": safe_str(cells[param_idx]),
                    "byte_pos": safe_int(cells[byte_idx]) if byte_idx and byte_idx < len(cells) else 0,
                    "bit_pos": safe_int(cells[bit_idx]) if bit_idx and bit_idx < len(cells) else 0,
                    "bit_length": safe_int(cells[bitlen_idx]) if bitlen_idx and bitlen_idx < len(cells) else 8,
                    "data_type": safe_str(cells[dt_idx]) if dt_idx and dt_idx < len(cells) else "RAW",
                }
                io_entry["parameters"].append(param)

            current_io = io_entry
            io_controls.append(io_entry)

        elif current_io is not None:
            # 参数续行
            param_idx = col_map.get("parameter")
            if param_idx is not None and param_idx < len(cells) and cells[param_idx] is not None:
                req_resp_idx = col_map.get("req_resp")
                byte_idx = col_map.get("byte_pos")
                bit_idx = col_map.get("bit_pos")
                bitlen_idx = col_map.get("bit_length")
                dt_idx = col_map.get("data_type")

                param = {
                    "req_resp": safe_str(cells[req_resp_idx]) if req_resp_idx and req_resp_idx < len(cells) else "Req",
                    "name": safe_str(cells[param_idx]),
                    "byte_pos": safe_int(cells[byte_idx]) if byte_idx and byte_idx < len(cells) else 0,
                    "bit_pos": safe_int(cells[bit_idx]) if bit_idx and bit_idx < len(cells) else 0,
                    "bit_length": safe_int(cells[bitlen_idx]) if bitlen_idx and bitlen_idx < len(cells) else 8,
                    "data_type": safe_str(cells[dt_idx]) if dt_idx and dt_idx < len(cells) else "RAW",
                }
                current_io["parameters"].append(param)

    return io_controls, defaults_used


def parse_routine_sheet(ws, sheet_name):
    """解析Routine Sheet"""
    routines = []
    defaults_used = []

    header_row_idx = find_header_row(ws, "routine")
    header_row = [safe_str(cell.value) for cell in list(ws.iter_rows(
        min_row=header_row_idx, max_row=header_row_idx, values_only=False))[0]]

    col_map = detect_column_mapping(header_row, "routine")

    current_routine = None

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=False):
        cells = [cell.value for cell in row]
        if any(safe_str(c).startswith("#End") for c in cells if c):
            break
        if all(c is None for c in cells):
            continue

        rid_idx = col_map.get("rid_number")
        rid_val = cells[rid_idx] if rid_idx is not None and rid_idx < len(cells) else None

        if rid_val is not None and safe_str(rid_val).strip():
            rid_hex = normalize_hex(rid_val)
            if rid_hex is None:
                continue

            ctrl_idx = col_map.get("control_type")
            routine_entry = {
                "rid_number": rid_hex,
                "rid_name": "",
                "rid_name_cn": "",
                "control_type": safe_str(cells[ctrl_idx]) if ctrl_idx and ctrl_idx < len(cells) and cells[ctrl_idx] else DEFAULTS["routine"]["control_type"],
                "conditions": DEFAULTS["routine"]["conditions"],
                "security_level": DEFAULTS["routine"]["security_level"],
                "req_parameters": [],
                "resp_parameters": [],
                "remarks": "",
            }

            for field in ["rid_name", "rid_name_cn", "conditions", "security_level", "remarks"]:
                idx = col_map.get(field)
                if idx is not None and idx < len(cells) and cells[idx] is not None:
                    routine_entry[field] = safe_str(cells[idx])

            routine_defaults = []
            if not routine_entry["conditions"] or routine_entry["conditions"] == DEFAULTS["routine"]["conditions"]:
                if col_map.get("conditions") is None:
                    routine_defaults.append(("conditions", DEFAULTS["routine"]["conditions"]))
            if not routine_entry["security_level"] or routine_entry["security_level"] == DEFAULTS["routine"]["security_level"]:
                if col_map.get("security_level") is None:
                    routine_defaults.append(("security_level", DEFAULTS["routine"]["security_level"]))

            if routine_defaults:
                defaults_used.append({"rid": rid_hex, "defaults": routine_defaults})

            # 解析参数
            req_resp_idx = col_map.get("req_resp")
            param_idx = col_map.get("parameter")
            byte_idx = col_map.get("byte_pos")
            bit_idx = col_map.get("bit_pos")
            bitlen_idx = col_map.get("bit_length")
            dt_idx = col_map.get("data_type")
            mt_idx = col_map.get("method_type")

            rr = safe_str(cells[req_resp_idx]).lower() if req_resp_idx and req_resp_idx < len(cells) and cells[req_resp_idx] else ""

            if param_idx is not None and param_idx < len(cells) and cells[param_idx] is not None:
                param = {
                    "name": safe_str(cells[param_idx]),
                    "byte_pos": safe_str(cells[byte_idx]) if byte_idx and byte_idx < len(cells) else "0",
                    "bit_pos": safe_str(cells[bit_idx]) if bit_idx and bit_idx < len(cells) else "0",
                    "bit_length": safe_int(cells[bitlen_idx], 8) if bitlen_idx and bitlen_idx < len(cells) else 8,
                    "data_type": safe_str(cells[dt_idx]) if dt_idx and dt_idx < len(cells) else "RAW",
                    "method_type": safe_str(cells[mt_idx]) if mt_idx and mt_idx < len(cells) else "",
                }
                if "resp" in rr:
                    routine_entry["resp_parameters"].append(param)
                else:
                    routine_entry["req_parameters"].append(param)

            current_routine = routine_entry
            routines.append(routine_entry)

        elif current_routine is not None:
            req_resp_idx = col_map.get("req_resp")
            param_idx = col_map.get("parameter")

            rr = safe_str(cells[req_resp_idx]).lower() if req_resp_idx and req_resp_idx < len(cells) and cells[req_resp_idx] else ""

            # 如果是新的Req/Resp行但无RID，说明是同一RID的响应行
            if "resp" in rr and not (param_idx and param_idx < len(cells) and cells[param_idx]):
                continue

            if param_idx is not None and param_idx < len(cells) and cells[param_idx] is not None:
                byte_idx = col_map.get("byte_pos")
                bit_idx = col_map.get("bit_pos")
                bitlen_idx = col_map.get("bit_length")
                dt_idx = col_map.get("data_type")
                mt_idx = col_map.get("method_type")

                param = {
                    "name": safe_str(cells[param_idx]),
                    "byte_pos": safe_str(cells[byte_idx]) if byte_idx and byte_idx < len(cells) else "0",
                    "bit_pos": safe_str(cells[bit_idx]) if bit_idx and bit_idx < len(cells) else "0",
                    "bit_length": safe_int(cells[bitlen_idx], 8) if bitlen_idx and bitlen_idx < len(cells) else 8,
                    "data_type": safe_str(cells[dt_idx]) if dt_idx and dt_idx < len(cells) else "RAW",
                    "method_type": safe_str(cells[mt_idx]) if mt_idx and mt_idx < len(cells) else "",
                }
                if "resp" in rr:
                    current_routine["resp_parameters"].append(param)
                else:
                    current_routine["req_parameters"].append(param)

    return routines, defaults_used


def parse_dtc_sheet(ws, sheet_name):
    """解析DTC Sheet"""
    dtcs = []
    defaults_used = []

    header_row_idx = find_header_row(ws, "dtc")
    header_row = [safe_str(cell.value) for cell in list(ws.iter_rows(
        min_row=header_row_idx, max_row=header_row_idx, values_only=False))[0]]

    col_map = detect_column_mapping(header_row, "dtc")

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=False):
        cells = [cell.value for cell in row]
        if any(safe_str(c).startswith("#End") for c in cells if c):
            break
        if all(c is None for c in cells):
            continue

        dtc_hex_idx = col_map.get("dtc_number_hex")
        dtc_val = cells[dtc_hex_idx] if dtc_hex_idx is not None and dtc_hex_idx < len(cells) else None

        if dtc_val is None:
            continue

        dtc_hex = normalize_hex(dtc_val)
        if dtc_hex is None:
            continue

        dtc_entry = {
            "dtc_number_hex": dtc_hex,
            "dtc_number": "",
            "dtc_name": "",
            "failure_type": DEFAULTS["dtc"]["failure_type"],
            "priority": DEFAULTS["dtc"]["priority"],
            "lamp_flag": DEFAULTS["dtc"]["lamp_flag"],
            "monitor_type": DEFAULTS["dtc"]["monitor_type"],
            "monitor_rate": DEFAULTS["dtc"]["monitor_rate"],
            "monitor_enable_criteria": "",
            "test_failed_criteria": "",
            "test_pass_criteria": "",
            "mature_time": "",
            "demature_time": "",
            "ecu_action": "",
            "dtc_aging": DEFAULTS["dtc"]["dtc_aging"],
            "snapshot": "",
        }

        field_map = {
            "dtc_number": "dtc_number", "dtc_name": "dtc_name",
            "failure_type": "failure_type", "lamp_flag": "lamp_flag",
            "monitor_type": "monitor_type", "monitor_rate": "monitor_rate",
            "monitor_enable": "monitor_enable_criteria",
            "test_failed_criteria": "test_failed_criteria",
            "test_pass_criteria": "test_pass_criteria",
            "mature_time": "mature_time", "demature_time": "demature_time",
            "ecu_action": "ecu_action", "snapshot": "snapshot",
        }

        for field, key in field_map.items():
            idx = col_map.get(field)
            if idx is not None and idx < len(cells) and cells[idx] is not None:
                dtc_entry[key] = safe_str(cells[idx])

        priority_idx = col_map.get("priority")
        if priority_idx is not None and priority_idx < len(cells):
            dtc_entry["priority"] = safe_int(cells[priority_idx], DEFAULTS["dtc"]["priority"])

        aging_idx = col_map.get("dtc_aging")
        if aging_idx is not None and aging_idx < len(cells):
            dtc_entry["dtc_aging"] = safe_int(cells[aging_idx], DEFAULTS["dtc"]["dtc_aging"])

        dtc_defaults = []
        if dtc_entry["failure_type"] == DEFAULTS["dtc"]["failure_type"] and col_map.get("failure_type") is None:
            dtc_defaults.append(("failure_type", DEFAULTS["dtc"]["failure_type"]))
        if dtc_entry["monitor_type"] == DEFAULTS["dtc"]["monitor_type"] and col_map.get("monitor_type") is None:
            dtc_defaults.append(("monitor_type", DEFAULTS["dtc"]["monitor_type"]))

        if dtc_defaults:
            defaults_used.append({"dtc": dtc_hex, "defaults": dtc_defaults})

        dtcs.append(dtc_entry)

    return dtcs, defaults_used


def extract_can_config(wb):
    """
    从Excel工作簿中提取CAN总线配置信息（tx_id, rx_id, bitrate, sample_point, can_fd等）。
    搜索所有sheet的前30行，查找包含这些关键字的单元格。
    返回 dict，只包含实际找到的字段。
    """
    config = {}
    patterns = {
        "tx_id": r"tx.?id|发送.?id|request.?id|diag.*tx|client.*id|tester.*id",
        "rx_id": r"rx.?id|接收.?id|response.?id|diag.*rx|ecu.*id|server.*id",
        "func_id": r"func.*id|功能.*id|functional.*id|broadcast.*id",
        "bitrate": r"\bbitrate\b|\bbaud.?rate\b|波特率|通信速率|bit\s*rate",
        "sample_point": r"sample.?point|采样点|sp\s*%|sample\s*%",
        "can_fd": r"\bcan\s*fd\b|\bcanfd\b|fd\s*mode|fd模式",
        "fd_data_bitrate": r"d(?:ata)?[\s_-]*(?:bit)?rate|数据段.*波特率|数据.*速率|dbitrate|data\s*baud",
        "fd_dsample_point": r"d(?:ata)?[\s_-]*sample|数据段.*采样|dsample|data\s*sample",
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=30, values_only=False):
            for cell_idx, cell in enumerate(row):
                label = safe_str(cell.value).lower().strip()
                if not label:
                    continue
                for field, pat in patterns.items():
                    if field in config:
                        continue
                    if re.search(pat, label, re.IGNORECASE):
                        if field in ("tx_id", "rx_id", "func_id"):
                            can_id = None
                            for raw in iter_nearby_values(ws, cell.row, cell.column):
                                can_id = normalize_can_id(raw)
                                if can_id is not None:
                                    config[field] = can_id
                                    break
                            if can_id is None:
                                print(f"[WARN] CAN配置字段 {field} 找到了标签，但附近未找到合法CAN ID，已忽略")
                        elif field in ("bitrate", "fd_data_bitrate"):
                            raw = None
                            for candidate in iter_nearby_values(ws, cell.row, cell.column):
                                raw = candidate
                                break
                            if raw is None:
                                continue
                            # 支持 "500kbps", "500000", "500K", "2Mbps", "2000000" 等
                            parsed_br = None
                            m = re.search(r'(\d+(?:\.\d+)?)\s*[mM]', raw)
                            if m:
                                parsed_br = int(float(m.group(1)) * 1000000)
                            else:
                                m = re.search(r'(\d+)\s*[kK]', raw)
                                if m:
                                    parsed_br = int(m.group(1)) * 1000
                                else:
                                    v = safe_int(raw, 0)
                                    if v > 0:
                                        parsed_br = v
                            # 合理范围校验: 10kbps ~ 8Mbps
                            if parsed_br is not None:
                                if 10000 <= parsed_br <= 8000000:
                                    config[field] = parsed_br
                                else:
                                    print(f"[WARN] CAN配置字段 {field} 的值 {parsed_br} 超出合理范围(10K~8M)，已忽略")
                        elif field in ("sample_point", "fd_dsample_point"):
                            raw = None
                            for candidate in iter_nearby_values(ws, cell.row, cell.column):
                                raw = candidate
                                break
                            if raw is None:
                                continue
                            # 支持 "80%", "0.8", "80", "75%" 等
                            m = re.search(r'(\d+(?:\.\d+)?)\s*%?', raw)
                            if m:
                                sp = float(m.group(1))
                                if sp > 1.0:
                                    sp = sp / 100.0  # "80" → 0.8
                                if 0.5 <= sp <= 0.95:
                                    config[field] = sp
                                else:
                                    print(f"[WARN] CAN配置字段 {field} 的采样点值 {sp:.3f} 超出合理范围(0.5~0.95)，已忽略")
                        elif field == "can_fd":
                            raw = None
                            for candidate in iter_nearby_values(ws, cell.row, cell.column):
                                raw = candidate
                                break
                            if raw is None:
                                continue
                            # 支持 "Yes", "Y", "True", "启用", "ON", "CAN FD" 等
                            lower = raw.lower()
                            if any(kw in lower for kw in ("yes", "y", "true", "启用", "on", "enable", "fd")):
                                config[field] = True
                            elif any(kw in lower for kw in ("no", "n", "false", "禁用", "off", "disable", "classic")):
                                config[field] = False
    return config


def parse_excel(filepath):
    """解析Excel格式的诊断调查表"""
    import openpyxl

    wb = openpyxl.load_workbook(filepath, data_only=True)
    result = {
        "source_file": os.path.basename(filepath),
        "can_config": {},
        "dids": [],
        "io_controls": [],
        "routines": [],
        "dtcs": [],
        "defaults_used": [],
        "missing_attributes": [],
    }

    # 提取CAN总线配置
    result["can_config"] = extract_can_config(wb)
    if result["can_config"]:
        print(f"[INFO] 从调查表中提取到CAN配置: {result['can_config']}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_type = detect_sheet_type(sheet_name, ws)

        if sheet_type is None:
            continue

        if sheet_type == "did":
            dids, defaults = parse_did_sheet(ws, sheet_name)
            result["dids"].extend(dids)
            result["defaults_used"].extend(defaults)
        elif sheet_type == "io_control":
            ios, defaults = parse_io_control_sheet(ws, sheet_name)
            result["io_controls"].extend(ios)
            result["defaults_used"].extend(defaults)
        elif sheet_type == "routine":
            routines, defaults = parse_routine_sheet(ws, sheet_name)
            result["routines"].extend(routines)
            result["defaults_used"].extend(defaults)
        elif sheet_type == "dtc":
            dtcs, defaults = parse_dtc_sheet(ws, sheet_name)
            result["dtcs"].extend(dtcs)
            result["defaults_used"].extend(defaults)

    # 去重合并DID: 同一DID号出现多次时保留会话信息最完整的条目
    if result["dids"]:
        seen = {}  # did_number → (index, completeness_score)
        merged = []
        for did in result["dids"]:
            dn = did["did_number"]
            # 计算完整度: 有多少session为"Y"
            score = sum(1 for v in did.get("read_sessions", {}).values() if v == "Y")
            score += sum(1 for v in did.get("write_sessions", {}).values() if v == "Y")
            score += sum(1 for v in did.get("read_security", {}).values() if v == "Y")
            score += sum(1 for v in did.get("write_security", {}).values() if v == "Y")
            if dn in seen:
                old_idx, old_score = seen[dn]
                if score > old_score:
                    merged[old_idx] = did
                    seen[dn] = (old_idx, score)
                # else keep old one
            else:
                seen[dn] = (len(merged), score)
                merged.append(did)
        result["dids"] = merged

    wb.close()
    return result


def parse_csv(filepath):
    """解析CSV格式的诊断调查表"""
    import csv

    result = {
        "source_file": os.path.basename(filepath),
        "can_config": {},
        "dids": [],
        "io_controls": [],
        "routines": [],
        "dtcs": [],
        "defaults_used": [],
        "missing_attributes": [],
    }

    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        return result

    # 简单启发式：根据表头判断类型
    header = " ".join(rows[0]).lower()
    if "dtc" in header:
        sheet_type = "dtc"
    elif "iocontrol" in header or "io control" in header:
        sheet_type = "io_control"
    elif "routine" in header or "rid" in header:
        sheet_type = "routine"
    else:
        sheet_type = "did"

    # 转换为openpyxl风格的工作表（简单模拟）
    # 对CSV使用简化的解析逻辑
    col_map = detect_column_mapping(rows[0], sheet_type)

    if sheet_type == "did":
        for row_data in rows[1:]:
            if not row_data or all(not c.strip() for c in row_data):
                continue
            did_idx = col_map.get("did_number")
            if did_idx is None or did_idx >= len(row_data):
                continue
            did_hex = normalize_hex(row_data[did_idx])
            if did_hex is None:
                continue

            did_entry = {
                "did_number": did_hex,
                "did_name": row_data[col_map["did_name"]] if "did_name" in col_map and col_map["did_name"] < len(row_data) else DEFAULTS["did"]["did_name"],
                "rw_state": row_data[col_map["rw_state"]] if "rw_state" in col_map and col_map["rw_state"] < len(row_data) else DEFAULTS["did"]["rw_state"],
                "size_bytes": safe_int(row_data[col_map["size_bytes"]]) if "size_bytes" in col_map and col_map["size_bytes"] < len(row_data) else DEFAULTS["did"]["size_bytes"],
                "data_type": row_data[col_map["data_type"]] if "data_type" in col_map and col_map["data_type"] < len(row_data) else DEFAULTS["did"]["data_type"],
                "sub_data": [],
                "read_sessions": {"default_0x01": "Y", "programming_0x02": "N", "extended_0x03": "Y"},
                "read_security": {"level0_locked": "Y", "level1": "N", "level_fbl": "N", "level_immo": "N"},
                "write_sessions": {"default_0x01": "N", "programming_0x02": "N", "extended_0x03": "N"},
                "write_security": {"level0_locked": "N", "level1": "N", "level_fbl": "N", "level_immo": "N"},
            }
            result["dids"].append(did_entry)

    return result


def parse_json(filepath):
    """解析JSON格式的诊断调查表"""
    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)

    # 如果已经是标准格式，直接返回
    if "dids" in data and "dtcs" in data:
        data.setdefault("can_config", {})
        data.setdefault("defaults_used", [])
        data.setdefault("missing_attributes", [])
        return data

    # 否则尝试转换
    result = {
        "source_file": os.path.basename(filepath),
        "can_config": data.get("can_config", {}),
        "dids": data.get("dids", []),
        "io_controls": data.get("io_controls", []),
        "routines": data.get("routines", []),
        "dtcs": data.get("dtcs", []),
        "defaults_used": [],
        "missing_attributes": [],
    }
    return result


def parse_survey_table(filepath):
    """根据文件类型自动选择解析器"""
    ext = Path(filepath).suffix.lower()

    if ext in (".xlsx", ".xls"):
        return parse_excel(filepath)
    elif ext == ".csv":
        return parse_csv(filepath)
    elif ext == ".json":
        return parse_json(filepath)
    else:
        print(f"[ERROR] 不支持的文件格式: {ext}")
        print("支持的格式: .xlsx, .xls, .csv, .json")
        sys.exit(1)


def main():
    parser = argparse.ArgumentParser(description="UDS诊断调查表解析器")
    parser.add_argument("--input", "-i", required=True, help="诊断调查表文件路径")
    parser.add_argument("--output", "-o", required=True, help="输出JSON文件路径")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"[ERROR] 文件不存在: {args.input}")
        sys.exit(1)

    print(f"[INFO] 解析诊断调查表: {args.input}")
    result = parse_survey_table(args.input)

    # 统计
    print(f"[INFO] 解析完成:")
    print(f"  - DID数量: {len(result['dids'])}")
    print(f"  - IOControl数量: {len(result['io_controls'])}")
    print(f"  - Routine数量: {len(result['routines'])}")
    print(f"  - DTC数量: {len(result['dtcs'])}")
    print(f"  - 使用默认值: {len(result['defaults_used'])} 项")

    if result["defaults_used"]:
        print("\n[WARNING] 以下属性使用了默认值，建议用户补充：")
        for item in result["defaults_used"]:
            ident = item.get("did") or item.get("io_did") or item.get("rid") or item.get("dtc") or "?"
            for attr, val in item.get("defaults", []):
                print(f"  - {ident}: {attr} = {val}")

    os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2, default=str)

    print(f"\n[INFO] 输出文件: {args.output}")


if __name__ == "__main__":
    main()
